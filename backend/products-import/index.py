"""
Импорт товаров из CSV/Excel с маппингом колонок от клиента.

Режимы:
  action=preview  — распарсить файл, вернуть заголовки + первые 3 строки данных
  action=import   — импорт с маппингом { file_col: db_field } и проверкой дублей
"""
import base64
import io
import json
import os

import psycopg2
from psycopg2.extras import RealDictCursor

SCHEMA = "t_p54266347_multi_level_electron"
CORS = {
    "Access-Control-Allow-Origin": "*",
    "Access-Control-Allow-Methods": "POST, OPTIONS",
    "Access-Control-Allow-Headers": "Content-Type",
}

NUMERIC_FIELDS = {
    "zakup_tsena_yuan", "kurs_yuan", "tsena_dostavki", "ves_tovara",
    "kurs_dollara", "stavka_kg", "stavka_kub", "sebestoimost",
    "fifo", "lifo", "prodazh_tsena_roznitsa", "prodazh_tsena_opt",
    "kolichestvo", "ostatok", "summa", "zakazano", "otgruzheno",
    "vozvrat_postavshchiku", "vozvrat_ot_pokupatelya",
}

# Все доступные поля БД с человекочитаемыми названиями
DB_FIELDS = [
    ("naimenovanie",           "Наименование *"),
    ("kod_kitay",              "Код товара Китай"),
    ("artikul",                "Артикул"),
    ("shtrikhkod",             "Штрихкод"),
    ("tsvet",                  "Цвет"),
    ("zakup_tsena_yuan",       "Закупочная цена (¥)"),
    ("kurs_yuan",              "Курс юаня"),
    ("tsena_dostavki",         "Цена доставки (₽)"),
    ("ves_tovara",             "Вес товара (кг)"),
    ("gabarity_upakovki",      "Габариты упаковки"),
    ("kurs_dollara",           "Курс доллара"),
    ("stavka_kg",              "Ставка за кг"),
    ("stavka_kub",             "Ставка за куб"),
    ("sebestoimost",           "Себестоимость (₽)"),
    ("fifo",                   "ФИФО (₽)"),
    ("lifo",                   "ЛИФО (₽)"),
    ("prodazh_tsena_roznitsa", "Цена розница (₽)"),
    ("prodazh_tsena_opt",      "Цена опт (₽)"),
    ("kolichestvo",            "Количество"),
    ("ostatok",                "Остаток"),
    ("summa",                  "Сумма (₽)"),
    ("zakazano",               "Заказано"),
    ("otgruzheno",             "Отгружено"),
    ("vozvrat_postavshchiku",  "Возврат поставщику"),
    ("vozvrat_ot_pokupatelya", "Возврат от покупателя"),
]

# Авто-маппинг по популярным названиям заголовков → db_field
AUTO_MAP = {
    "наименование": "naimenovanie", "название": "naimenovanie", "наимен": "naimenovanie",
    "код китай": "kod_kitay", "код товара китай": "kod_kitay", "code": "kod_kitay",
    "артикул": "artikul", "арт": "artikul", "sku": "artikul",
    "штрихкод": "shtrikhkod", "barcode": "shtrikhkod", "ean": "shtrikhkod",
    "цвет": "tsvet", "color": "tsvet",
    "закупочная цена юань": "zakup_tsena_yuan", "цена юань": "zakup_tsena_yuan", "цена ¥": "zakup_tsena_yuan",
    "курс юаня": "kurs_yuan", "курс ю": "kurs_yuan",
    "цена доставки": "tsena_dostavki", "доставка": "tsena_dostavki",
    "вес товара": "ves_tovara", "вес": "ves_tovara", "weight": "ves_tovara",
    "габариты упаковки": "gabarity_upakovki", "габариты": "gabarity_upakovki",
    "курс доллара": "kurs_dollara", "курс $": "kurs_dollara",
    "ставка кг": "stavka_kg", "ставка за кг": "stavka_kg",
    "ставка куб": "stavka_kub", "ставка за куб": "stavka_kub",
    "себестоимость": "sebestoimost",
    "фифо": "fifo", "fifo": "fifo",
    "лифо": "lifo", "lifo": "lifo",
    "продажная розница цена": "prodazh_tsena_roznitsa", "цена розница": "prodazh_tsena_roznitsa", "розница": "prodazh_tsena_roznitsa",
    "продажная цена опт": "prodazh_tsena_opt", "цена опт": "prodazh_tsena_opt", "опт": "prodazh_tsena_opt",
    "количество": "kolichestvo", "кол-во": "kolichestvo", "qty": "kolichestvo",
    "остаток": "ostatok", "stock": "ostatok",
    "сумма": "summa", "total": "summa",
    "заказано": "zakazano",
    "отгружено": "otgruzheno",
    "возврат поставщику": "vozvrat_postavshchiku",
    "возврат от покупателя": "vozvrat_ot_pokupatelya",
}


def resp(status, body):
    return {
        "statusCode": status,
        "headers": {**CORS, "Content-Type": "application/json"},
        "body": json.dumps(body, ensure_ascii=False, default=str),
    }


def parse_file(file_bytes: bytes, filename: str):
    """Возвращает (headers_raw: list[str], rows: list[list[str]])"""
    fn = filename.lower()
    if fn.endswith(".xlsx"):
        return parse_xlsx(file_bytes)
    if fn.endswith(".xls"):
        return parse_xls(file_bytes)
    # CSV/TXT
    content = ""
    for enc in ("utf-8-sig", "utf-8", "cp1251", "latin-1"):
        try:
            content = file_bytes.decode(enc)
            break
        except Exception:
            pass
    return parse_csv(content)


def parse_csv(content: str):
    lines = [l for l in content.splitlines() if l.strip()]
    if not lines:
        return [], []
    sep = ";" if lines[0].count(";") >= lines[0].count(",") else ","

    def split_row(line):
        result, current, in_q = [], "", False
        for ch in line:
            if ch == '"':
                in_q = not in_q
            elif ch == sep and not in_q:
                result.append(current.strip().strip('"'))
                current = ""
            else:
                current += ch
        result.append(current.strip().strip('"'))
        return result

    headers = split_row(lines[0])
    rows = [split_row(l) for l in lines[1:] if l.strip()]
    return headers, rows


def parse_xlsx(data: bytes):
    """Парсит .xlsx через openpyxl — надёжно, с учётом пустых ячеек."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    all_rows = []
    for row in ws.iter_rows(values_only=True):
        vals = [str(v).strip() if v is not None else "" for v in row]
        all_rows.append(vals)
    wb.close()
    if not all_rows:
        return [], []
    # Нормализуем ширину
    max_w = max(len(r) for r in all_rows)
    all_rows = [r + [""] * (max_w - len(r)) for r in all_rows]
    headers = all_rows[0]
    data_rows = [r for r in all_rows[1:] if any(v for v in r)]
    return headers, data_rows


def parse_xls(data: bytes):
    """Парсит старый формат .xls через xlrd."""
    import xlrd
    wb = xlrd.open_workbook(file_contents=data)
    ws = wb.sheet_by_index(0)
    if ws.nrows == 0:
        return [], []
    all_rows = []
    for i in range(ws.nrows):
        row = []
        for j in range(ws.ncols):
            cell = ws.cell(i, j)
            # xlrd типы: 0=empty,1=text,2=number,3=date,4=bool,5=error
            if cell.ctype == 2 and cell.value == int(cell.value):
                row.append(str(int(cell.value)))
            elif cell.ctype in (0, 5):
                row.append("")
            else:
                row.append(str(cell.value).strip())
        all_rows.append(row)
    headers = all_rows[0]
    data_rows = [r for r in all_rows[1:] if any(v for v in r)]
    return headers, data_rows


def handler(event: dict, context) -> dict:
    if event.get("httpMethod") == "OPTIONS":
        return {"statusCode": 200, "headers": CORS, "body": ""}
    if event.get("httpMethod") != "POST":
        return resp(405, {"error": "Только POST"})

    body = json.loads(event.get("body") or "{}")
    action = body.get("action", "preview")
    file_b64 = body.get("file_base64", "")
    filename = body.get("filename", "file.csv").lower()

    if not file_b64:
        return resp(400, {"error": "Нет файла"})

    file_bytes = base64.b64decode(file_b64)
    headers_raw, data_rows = parse_file(file_bytes, filename)

    if not headers_raw:
        return resp(400, {"error": "Файл пустой или не удалось прочитать заголовки"})

    # ── PREVIEW ──────────────────────────────────────────────────────────────
    if action == "preview":
        # Авто-маппинг: угадываем соответствия
        auto = {}
        for i, h in enumerate(headers_raw):
            db = AUTO_MAP.get(h.strip().lower())
            if db:
                auto[h] = db

        # Первые 3 строки для предпросмотра
        preview_rows = []
        for row in data_rows[:3]:
            padded = row + [""] * max(0, len(headers_raw) - len(row))
            preview_rows.append(padded[:len(headers_raw)])

        return resp(200, {
            "headers": headers_raw,
            "preview_rows": preview_rows,
            "auto_mapping": auto,
            "db_fields": [{"key": k, "label": l} for k, l in DB_FIELDS],
            "total_rows": len(data_rows),
        })

    # ── IMPORT ───────────────────────────────────────────────────────────────
    if action == "import":
        # mapping: { "Колонка файла": "db_field" | "" }
        mapping = body.get("mapping", {})
        # Только активные маппинги (не пустые)
        active = {col: db for col, db in mapping.items() if db}

        if "naimenovanie" not in active.values():
            return resp(400, {"error": "Необходимо привязать колонку к полю «Наименование»"})

        conn = psycopg2.connect(os.environ["DATABASE_URL"])
        cur = conn.cursor(cursor_factory=RealDictCursor)

        added, skipped, errors = 0, [], []

        for row in data_rows:
            padded = row + [""] * max(0, len(headers_raw) - len(row))
            # Собираем запись по маппингу
            item = {}
            for col, db_field in active.items():
                try:
                    col_idx = headers_raw.index(col)
                    val = padded[col_idx].strip() if col_idx < len(padded) else ""
                    if val:
                        item[db_field] = val
                except ValueError:
                    pass

            name = item.get("naimenovanie", "").strip()
            if not name:
                errors.append("Пропущена строка без наименования")
                continue

            # Проверка дубля
            cur.execute(
                f"SELECT id FROM {SCHEMA}.products WHERE LOWER(naimenovanie) = LOWER(%s)",
                (name,)
            )
            if cur.fetchone():
                skipped.append(name)
                continue

            # Конвертируем числовые поля
            for f in NUMERIC_FIELDS:
                if f in item and item[f] != "":
                    try:
                        item[f] = float(str(item[f]).replace(",", ".").replace(" ", "").replace("\xa0", ""))
                    except Exception:
                        item[f] = 0

            clean = {k: v for k, v in item.items() if v != "" and v is not None}
            if not clean:
                continue

            cols_sql = ", ".join(clean.keys())
            placeholders = ", ".join(["%s"] * len(clean))
            cur.execute(
                f"INSERT INTO {SCHEMA}.products ({cols_sql}) VALUES ({placeholders})",
                list(clean.values())
            )
            added += 1

        conn.commit()
        cur.close()
        conn.close()

        return resp(200, {
            "added": added,
            "skipped": skipped,
            "skipped_count": len(skipped),
            "errors": errors,
            "total_in_file": len(data_rows),
        })

    return resp(400, {"error": f"Неизвестный action: {action}"})